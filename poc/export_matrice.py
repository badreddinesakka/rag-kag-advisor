# -*- coding: utf-8 -*-
"""
export_matrice.py — produit la matrice de conformité au format demandé par l'appel d'offres.

Le RFP (section 5.3) impose une réponse CLAUSE PAR CLAUSE, avec une colonne de
conformité ne pouvant prendre qu'une seule de trois valeurs : Fully compliant,
Partially compliant, Not compliant. Ce fichier reproduit exactement cette
structure, une ligne par exigence extraite.

C'est le livrable attendu — pas un résumé de l'appel d'offres, mais le tableau
que le fournisseur remplira.

Utilisation :
    python export_matrice.py
    python export_matrice.py --entree exigences_extraites.json --sortie matrice.xlsx
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

POLICE = "Arial"

# Les trois seules valeurs acceptées par le RFP (section 5.3.2).
CONFORMITE = ["Fully compliant", "Partially compliant", "Not compliant"]

ROUGE_OOREDOO = "C8102E"
GRIS_CLAIR    = "F2F2F2"
JAUNE         = "FFF2CC"   # cellules que le fournisseur doit remplir

COLONNES = [
    ("ID", 10),
    ("Categorie", 14),
    ("Exigence", 62),
    ("Valeur / Seuil", 22),
    ("Statut", 14),
    ("Compliance", 20),
    ("Commentaire fournisseur", 34),
    ("Reference (doc + page)", 22),
    ("Source (morceau)", 26),
]


def _bordure() -> Border:
    fin = Side(style="thin", color="BFBFBF")
    return Border(left=fin, right=fin, top=fin, bottom=fin)


def construire(donnees: dict, sortie: str) -> str:
    exigences = donnees["exigences"]

    wb = Workbook()

    # ----- feuille 1 : la matrice --------------------------------------
    ws = wb.active
    ws.title = "Compliance Matrix"

    ws["A1"] = "Ooredoo Tunisia — Bill Generation Solution"
    ws["A1"].font = Font(name=POLICE, size=13, bold=True)
    ws["A2"] = ("Matrice de conformite — une ligne par exigence. "
                "Le fournisseur remplit les colonnes en jaune.")
    ws["A2"].font = Font(name=POLICE, size=9, italic=True)
    ws["A3"] = (f"Exigences extraites automatiquement : "
                f"{donnees['exigences_apres_fusion']} · "
                f"decoupage {donnees['child_tokens']}/{donnees['parent_tokens']} tokens "
                f"· modele {donnees['modele']}. "
                f"A RELIRE ET COMPLETER A LA MAIN avant envoi.")
    ws["A3"].font = Font(name=POLICE, size=9, color="C00000")

    ligne_entete = 5
    for i, (titre, largeur) in enumerate(COLONNES, start=1):
        cell = ws.cell(row=ligne_entete, column=i, value=titre)
        cell.font = Font(name=POLICE, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=ROUGE_OOREDOO)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _bordure()
        ws.column_dimensions[get_column_letter(i)].width = largeur

    for j, x in enumerate(exigences):
        r = ligne_entete + 1 + j
        valeurs = [
            x["id"],
            x["categorie"],
            x["exigence"],
            x["valeur"],
            x["statut"],
            "",                      # Compliance — a remplir
            "",                      # Commentaire — a remplir
            "",                      # Reference — a remplir
            "; ".join(x["sources"][:3]),
        ]
        for i, v in enumerate(valeurs, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = Font(name=POLICE, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(i in (3, 7)))
            cell.border = _bordure()
            if i in (6, 7, 8):
                cell.fill = PatternFill("solid", fgColor=JAUNE)
            elif j % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=GRIS_CLAIR)

    derniere = ligne_entete + len(exigences)

    # Liste deroulante sur la colonne Compliance : le RFP interdit toute autre
    # valeur que ces trois-la.
    if exigences:
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(CONFORMITE) + '"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Valeur non autorisee",
            error="Le RFP (5.3.2) n'accepte que : "
                  "Fully compliant, Partially compliant, Not compliant.",
        )
        ws.add_data_validation(dv)
        dv.add(f"F{ligne_entete + 1}:F{derniere}")

    ws.freeze_panes = f"A{ligne_entete + 1}"
    ws.auto_filter.ref = f"A{ligne_entete}:I{derniere}"

    # ----- feuille 2 : le recapitulatif ---------------------------------
    ws2 = wb.create_sheet("Recapitulatif")
    ws2["A1"] = "Recapitulatif"
    ws2["A1"].font = Font(name=POLICE, size=13, bold=True)

    r = 3
    ws2.cell(row=r, column=1, value="Total exigences").font = Font(
        name=POLICE, size=10, bold=True)
    ws2.cell(row=r, column=2,
             value=f"=COUNTA('Compliance Matrix'!A{ligne_entete + 1}:"
                   f"A{derniere})")
    r += 2

    for titre, dico in (("Par categorie", donnees["par_categorie"]),
                        ("Par statut", donnees["par_statut"])):
        ws2.cell(row=r, column=1, value=titre).font = Font(
            name=POLICE, size=11, bold=True)
        r += 1
        for cle in dico:
            ws2.cell(row=r, column=1, value=cle).font = Font(name=POLICE, size=10)
            # Formule, pas une valeur figée : le compte suit les modifications
            # faites à la main dans la matrice.
            ws2.cell(row=r, column=2,
                     value=f'=COUNTIF(\'Compliance Matrix\'!'
                           f'{"B" if titre.endswith("categorie") else "E"}'
                           f'{ligne_entete + 1}:'
                           f'{"B" if titre.endswith("categorie") else "E"}'
                           f'{derniere},A{r})')
            r += 1
        r += 1

    ws2.cell(row=r, column=1, value="Avancement du remplissage").font = Font(
        name=POLICE, size=11, bold=True)
    r += 1
    for etiquette in CONFORMITE:
        ws2.cell(row=r, column=1, value=etiquette).font = Font(name=POLICE, size=10)
        ws2.cell(row=r, column=2,
                 value=f"=COUNTIF('Compliance Matrix'!F{ligne_entete + 1}:"
                       f"F{derniere},A{r})")
        r += 1
    ws2.cell(row=r, column=1, value="Non renseigne").font = Font(
        name=POLICE, size=10, bold=True)
    ws2.cell(row=r, column=2,
             value=f"=COUNTBLANK('Compliance Matrix'!F{ligne_entete + 1}:"
                   f"F{derniere})")

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14

    # ----- feuille 3 : mode d'emploi ------------------------------------
    ws3 = wb.create_sheet("Mode d'emploi")
    lignes = [
        ("Comment utiliser ce fichier", True),
        ("", False),
        ("Colonnes en JAUNE : a remplir par le fournisseur.", False),
        ("  Compliance   : liste deroulante, 3 valeurs imposees par le RFP (5.3.2).", False),
        ("  Commentaire  : justification libre.", False),
        ("  Reference    : document + numero de page, exige par le RFP.", False),
        ("", False),
        ("Colonnes blanches : extraites automatiquement du document.", False),
        ("  Statut       : deduit du texte (must/shall = obligatoire,", False),
        ("                 should/recommended = recommande, sinon non precise).", False),
        ("  Source       : le morceau du document d'ou vient l'exigence,", False),
        ("                 pour pouvoir verifier.", False),
        ("", False),
        ("AVERTISSEMENT", True),
        ("Cette matrice est produite par un modele de langage local. Elle est un", False),
        ("POINT DE DEPART, pas un livrable. Chaque ligne doit etre relue :", False),
        ("  - une exigence peut avoir ete manquee ;", False),
        ("  - une ligne peut etre un fragment de phrase et non une exigence ;", False),
        ("  - un statut peut avoir ete mal deduit.", False),
        ("", False),
        ("Le RFP precise que toute clause marquee 'Fully compliant' ENGAGE le", False),
        ("fournisseur pour toute la duree du contrat (section 5.3.2).", False),
    ]
    for i, (texte, gras) in enumerate(lignes, start=1):
        c = ws3.cell(row=i, column=1, value=texte)
        c.font = Font(name=POLICE, size=11 if gras else 10, bold=gras)
    ws3.column_dimensions["A"].width = 90

    wb.save(sortie)
    return sortie


def main():
    ap = argparse.ArgumentParser(
        description="Exporte les exigences en matrice de conformite Excel.")
    ap.add_argument("--entree", default="exigences_extraites.json")
    ap.add_argument("--sortie", default="OT_Compliance_Matrix_auto.xlsx")
    args = ap.parse_args()

    chemin = Path(args.entree)
    if not chemin.exists():
        raise SystemExit(f"{args.entree} introuvable. "
                         f"Lance d'abord : python extract_criteres.py --input dossier_rfp")

    donnees = json.loads(chemin.read_text(encoding="utf-8"))
    fichier = construire(donnees, args.sortie)
    print(f"[OK] {donnees['exigences_apres_fusion']} exigences écrites dans {fichier}")
    print("     Feuille 1 : la matrice · Feuille 2 : récapitulatif · "
          "Feuille 3 : mode d'emploi")


if __name__ == "__main__":
    main()
